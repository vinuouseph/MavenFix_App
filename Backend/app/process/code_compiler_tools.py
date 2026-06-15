import subprocess
import os
import platform
import uuid
from xhtml2pdf import pisa


def compile_spring_boot(project_path, build_tool="maven"):
    """
    Compiles a Spring Boot project using Python.

    :param project_path: The absolute or relative path to the Spring Boot project.
    :param build_tool: 'maven' or 'gradle'
    """

    if not os.path.exists(project_path):
        msg = f"Error: Directory '{project_path}' does not exist."
        print(msg)
        return False, msg

    # Determine the correct wrapper command based on the OS
    is_windows = platform.system() == "Windows"

    if build_tool.lower() == "maven":
        if is_windows:
            wrapper = "mvnw.cmd"
        else:
            wrapper = "./mvnw"
            
        if os.path.isfile(os.path.join(project_path, wrapper.replace("./", ""))):
            command = [wrapper, "clean", "package"]
        else:
            command = ["mvn", "clean", "package"]
            
    elif build_tool.lower() == "gradle":
        if is_windows:
            wrapper = "gradlew.bat"
        else:
            wrapper = "./gradlew"
            
        if os.path.isfile(os.path.join(project_path, wrapper.replace("./", ""))):
            command = [wrapper, "build"]
        else:
            command = ["gradle", "build"]
    else:
        msg = "Error: build_tool must be 'maven' or 'gradle'."
        print(msg)
        return False, msg

    print(f"Starting compilation using {build_tool.capitalize()} in {project_path}...")

    try:
        result = subprocess.run(
            command,
            cwd=project_path,
            check=True,
            capture_output=True,
            text=True
        )

        print("✅ Compilation Successful!")
        print("\n".join(result.stdout.splitlines()[-10:]))
        return True, "Compilation Successful!"

    except subprocess.CalledProcessError as e:
        print("❌ Compilation Failed!")
        print("\n--- Error Output ---")
        msg = e.stderr if e.stderr else e.stdout
        print(msg)
        return False, msg
    except FileNotFoundError:
        msg = f"Error: Could not find the wrapper ({command[0]}).\nMake sure you have the Maven/Gradle wrapper in your project root, or change the script to use 'mvn' or 'gradle' directly if they are installed globally."
        print(msg)
        return False, msg


def _split_diff_by_file(diff_text: str) -> list[dict]:
    """
    Splits a unified diff string into per-file chunks and merges all hunks
    for the same file into a single block (like git tracking).

    Handles both:
      - "diff --git a/... b/..." style (git diffs)
      - "--- a/..." / "+++ b/..." style (difflib.unified_diff output)

    Returns a list of dicts ordered by first appearance:
      [{"filename": str, "diff": str}, ...]
    """
    import re

    # Only split on file-level boundaries; do NOT split on --- lines because
    # those are part of each hunk header and appear multiple times per file.
    file_header_re = re.compile(
        r'^(?:diff --git a/.+ b/(.+)|---\s+a/(.+))$',
        re.MULTILINE
    )

    positions = []
    for m in file_header_re.finditer(diff_text):
        filename = (m.group(1) or m.group(2) or "unknown").strip()
        positions.append((m.start(), filename))

    if not positions:
        # No file headers found — return the whole thing as one block
        return [{"filename": "Changes", "diff": diff_text.strip()}]

    # Collect raw per-block text (one entry per header match)
    raw_blocks: list[dict] = []
    for i, (start, filename) in enumerate(positions):
        end = positions[i + 1][0] if i + 1 < len(positions) else len(diff_text)
        raw_blocks.append({"filename": filename, "raw": diff_text[start:end].strip()})

    # Merge blocks that belong to the same file into a single card.
    # Strategy: keep the file header lines (---, +++) from the *first* occurrence
    # of each file, then append only the @@ hunk lines from subsequent occurrences.
    seen: dict[str, int] = {}   # filename -> index in merged list
    merged: list[dict] = []

    hunk_re = re.compile(r'^@@.', re.MULTILINE)

    for block in raw_blocks:
        fname = block["filename"]
        raw   = block["raw"]

        if fname not in seen:
            seen[fname] = len(merged)
            merged.append({"filename": fname, "diff": raw})
        else:
            # Find where the first @@ hunk starts and append only from there
            m = hunk_re.search(raw)
            if m:
                extra_hunks = raw[m.start():].strip()
                merged[seen[fname]]["diff"] += "\n" + extra_hunks

    return merged


def _render_diff_lines_html(diff_text: str) -> str:
    """
    Converts a raw diff string to HTML with per-line colouring as a single
    contiguous block.  Uses a collapsed <table> so xhtml2pdf never inserts
    the visible line-separators that appear with display:block <span> elements.

    Meta-lines (---, +++) are suppressed since the filename is already shown
    in the dark card header above.
    """
    rows = []
    for raw_line in diff_text.splitlines():
        # Suppress redundant file-path header lines — already shown in card header
        if (raw_line.startswith("---") or raw_line.startswith("+++")
                or raw_line.startswith("diff ") or raw_line.startswith("index ")):
            continue

        escaped = (
            raw_line
            .replace("&", "&amp;")
            .replace("<", "&lt;")
            .replace(">", "&gt;")
        ) or "&nbsp;"

        if raw_line.startswith("+") and not raw_line.startswith("+++"):
            bg, fg, gutter, fw = "#f0fdf4", "#15803d", "#22c55e", "normal"
        elif raw_line.startswith("-") and not raw_line.startswith("---"):
            bg, fg, gutter, fw = "#fff1f2", "#b91c1c", "#f87171", "normal"
        elif raw_line.startswith("@@"):
            bg, fg, gutter, fw = "#eff6ff", "#1d4ed8", "#93c5fd", "bold"
        else:
            bg, fg, gutter, fw = "#ffffff", "#1e293b", "#e2e8f0", "normal"

        rows.append(
            f'<tr>'
            f'<td style="width:3px;background:{gutter};padding:0;border:none;"></td>'
            f'<td style="background:{bg};color:{fg};font-weight:{fw};'
            f'padding:1px 10px 1px 8px;border:none;white-space:pre-wrap;'
            f'word-break:break-all;">'
            f'{escaped}'
            f'</td></tr>'
        )

    rows_html = "\n".join(rows)
    return (
        '<table style="width:100%;border-collapse:collapse;'
        'border-spacing:0;border:none;font-family:\'Courier New\',Courier,monospace;'
        'font-size:10.5px;line-height:1.65;">'
        f'{rows_html}'
        '</table>'
    )


def create_pdf_for_changes(md_content: str):
    clean_md = md_content.replace("→", "->")

    # Split into per-file chunks
    file_chunks = _split_diff_by_file(clean_md)

    # Build one card block per file
    file_cards_html = []
    for idx, chunk in enumerate(file_chunks, start=1):
        filename = chunk["filename"]
        diff_html = _render_diff_lines_html(chunk["diff"])
        card = f"""
        <div class="file-card">
            <div class="file-header">
                <span class="file-icon">&#x1F4C4;</span>
                <span class="file-name">{filename}</span>
            </div>
            <div class="diff-block">
                {diff_html}
            </div>
        </div>
        """
        file_cards_html.append(card)

    all_cards = "\n".join(file_cards_html)
    total_files = len(file_chunks)

    html_content = f"""<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <style>
        @page {{
            margin: 1.8cm 1.5cm;
        }}
        body {{
            font-family: Helvetica, Arial, sans-serif;
            color: #18181b;
            font-size: 12px;
            line-height: 1.5;
            background: #f3f4f6;
        }}
        .report-title {{
            font-size: 22px;
            font-weight: bold;
            color: #0f172a;
            margin-bottom: 4px;
        }}
        .report-meta {{
            font-size: 11px;
            color: #64748b;
            margin-bottom: 24px;
            padding-bottom: 10px;
            border-bottom: 2px solid #e2e8f0;
        }}
        /* ── Code-block card ── */
        .file-card {{
            background: #ffffff;
            border: 1px solid #d1d5db;
            border-radius: 8px;
            margin-bottom: 22px;
            overflow: hidden;
            page-break-inside: avoid;
        }}
        /* Dark header bar — exactly like ChatGPT / Claude code blocks */
        .file-header {{
            background: #1e1e2e;
            color: #cdd6f4;
            padding: 7px 14px;
            font-size: 11px;
            font-family: "Courier New", Courier, monospace;
            border-bottom: 1px solid #313244;
        }}
        .file-icon {{
            margin-right: 7px;
        }}
        .file-name {{
            color: #89dceb;
            font-size: 11px;
            word-break: break-all;
        }}
        /* Diff body — flush against the header */
        .diff-block {{
            background: #ffffff;
            font-family: "Courier New", Courier, monospace;
            font-size: 10.5px;
            line-height: 1.6;
        }}
    </style>
</head>
<body>
    <div class="report-title">Code Change Report</div>
    <div class="report-meta">
        Project Fix Agent &nbsp;|&nbsp; {total_files} file(s) modified
    </div>
    {all_cards}
</body>
</html>"""

    source_filename = "code_change"
    os.makedirs("exports", exist_ok=True)
    base = "".join(char if char.isalnum() or char in {"-", "_"} else "_" for char in source_filename)
    filename = f"{base}_{uuid.uuid4().hex[:8]}.pdf"
    path = os.path.join("exports", filename)

    with open(path, "wb") as pdf_file:
        pisa_status = pisa.CreatePDF(html_content, dest=pdf_file)

    if pisa_status.err:
        raise Exception(f"Failed to generate PDF. Errors: {pisa_status.err}")

    return filename



def update_metadata_excel(vuln_updates: list) -> str:
    """
    Updates the Metadata_Sheet.xlsx with the given vulnerability updates (POM dependencies).
    Copies the original excel with a UUID prefix, inserts the rows, and returns a download link.
    """
    if not vuln_updates:
        return ""

    import uuid
    from pathlib import Path
    import shutil
    import pandas as pd
    
    backend_dir = Path(__file__).resolve().parents[2]
    src_excel = backend_dir / "Metadata_Sheet.xlsx"
    
    if not src_excel.exists():
        print(f"Metadata_Sheet.xlsx not found at {src_excel}.")
        return ""
        
    uid = uuid.uuid4().hex[:8]
    dest_filename = f"{uid}_Metadata_Sheet.xlsx"
    
    exports_dir = backend_dir / "exports"
    exports_dir.mkdir(exist_ok=True)
    dest_excel = exports_dir / dest_filename
    
    shutil.copy2(src_excel, dest_excel)
    
    try:
        df = pd.read_excel(dest_excel)
        
        new_rows = []
        for update in vuln_updates:
            dep = update.get("dep", "")
            parts = dep.split(":")
            group_id = parts[0] if len(parts) > 0 else ""
            artifact_id = parts[1] if len(parts) > 1 else dep
            to_ver = update.get("to_ver", "")
            
            jar_name = f"{artifact_id}-{to_ver}.jar"
            download_url = f"https://repo1.maven.org/maven2/{group_id.replace('.', '/')}/{artifact_id}/{to_ver}/{jar_name}"
            
            row = {
                'Jar Name': jar_name,
                'Group Id': group_id,
                'Download URL': download_url,
                'License Type': '',
                'Packaging(Yes/No)': 'Yes',
                'Packaging in which deliveries': '',
                'Manifested(Yes/No)': 'Yes',
                'Java build versions': ''
            }
            new_rows.append(row)
            
        if new_rows:
            new_df = pd.DataFrame(new_rows)
            # Use pandas concat
            df = pd.concat([df, new_df], ignore_index=True)
            df.to_excel(dest_excel, index=False)
            
            # Apply yellow background to the header row and a modern font to the whole sheet
            import openpyxl
            from openpyxl.styles import PatternFill, Font
            
            wb = openpyxl.load_workbook(dest_excel)
            ws = wb.active
            
            yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            header_font = Font(name="Calibri", size=10, bold=True)
            body_font = Font(name="Calibri", size=10)
            
            # Apply body font to all cells
            for row in ws.iter_rows():
                for cell in row:
                    cell.font = body_font
            
            # Apply yellow fill and bold font to header
            for cell in ws[1]:
                cell.fill = yellow_fill
                cell.font = header_font
                
            wb.save(dest_excel)
            
        return dest_filename
    except Exception as e:
        print(f"Error updating metadata excel: {e}")
        return ""
async def generate_reason_for_changes_async(full_diff: str) -> str:
    """
    Calls the LLM asynchronously to explain the reasons for the code changes.
    """
    if not full_diff.strip():
        return "No changes were made."

    from app.llm.llm_registry import build_llm_model
    from langchain_core.messages import SystemMessage, HumanMessage

    llm = build_llm_model()
    
    sys_msg = SystemMessage(content="You are an expert software engineer. Your task is to analyze a unified diff of code changes and explain the clear, concise, and structured reasons behind these changes. Focus on the 'why'. Do not repeat the code, but explain the rationale for the modifications (e.g. 'Updated deprecated dependency to resolve vulnerability CVE-XXX', 'Fixed compiler error by importing missing class', 'Replaced javax with jakarta for Spring Boot 3 compatibility'). For every code change, you must also include a confidence score out of 100 representing how confident you are that the change is correct and optimal. You must wrap the confidence score in an HTML span tag like this: <span class=\"confidence-score\">[Confidence: 95/100]</span>. Format your output in Markdown with bullet points.")
    human_msg = HumanMessage(content=f"Here is the diff of the changes made:\n\n{full_diff}\n\nPlease provide the reasons for these changes, including the confidence score out of 100 for each, wrapped in the requested span tag.")
    
    try:
        response = await llm.ainvoke([sys_msg, human_msg])
        content = response.content
        if isinstance(content, list):
            text_parts = []
            for item in content:
                if isinstance(item, dict) and "text" in item:
                    text_parts.append(item["text"])
                elif isinstance(item, str):
                    text_parts.append(item)
                else:
                    text_parts.append(str(item))
            return "\n".join(text_parts)
        return str(content)
    except Exception as e:
        import logging
        logging.error(f"Failed to generate reason for changes: {e}")
        return "Failed to generate reason for changes due to an internal error."

def create_reason_pdf(reason_md: str) -> str:
    """
    Creates a PDF from the markdown reasoning and returns the filename.
    """
    import markdown
    import uuid
    import os
    from xhtml2pdf import pisa
    
    html_body = markdown.markdown(reason_md)
    html_content = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <style>
            @page {{ size: A4; margin: 2cm; }}
            body {{ font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, Helvetica, Arial, sans-serif; font-size: 14px; line-height: 1.6; color: #333; }}
            h1, h2, h3 {{ color: #2563eb; }}
            code {{ font-family: 'Courier New', Courier, monospace; background: #f1f5f9; padding: 2px 4px; border-radius: 4px; font-size: 13px; }}
            pre {{ background: #f8fafc; padding: 12px; border-radius: 6px; border: 1px solid #e2e8f0; }}
            ul, ol {{ margin-bottom: 16px; }}
            li {{ margin-bottom: 8px; }}
            .confidence-score {{ color: #d97706; font-weight: bold; background: #fef3c7; padding: 2px 6px; border-radius: 4px; font-size: 12px; }}
        </style>
    </head>
    <body>
        <h2>Reasoning for Code Changes</h2>
        {html_body}
    </body>
    </html>
    """
    
    os.makedirs("exports", exist_ok=True)
    filename = f"change_reasons_{uuid.uuid4().hex[:8]}.pdf"
    path = os.path.join("exports", filename)
    
    with open(path, "wb") as pdf_file:
        pisa_status = pisa.CreatePDF(html_content, dest=pdf_file)
        
    if pisa_status.err:
        raise Exception(f"Failed to generate reason PDF. Errors: {pisa_status.err}")
        
    return filename
